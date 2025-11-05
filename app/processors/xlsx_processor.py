from pathlib import Path
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook
from app.utils.logging_config import log
from app.utils.chunking import chunk_text


class XLSXProcessor:
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xlsm']
    
    def can_process(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() in self.supported_extensions
    
    def extract_text_from_xlsx(self, xlsx_path: str) -> str:
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            text = ""
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            
            wb.close()
            return text.strip()
        except Exception as e:
            log.error(f"Error extracting text from XLSX: {e}")
            return ""
    
    def process(self, file_path: str) -> Tuple[List[str], Dict[str, Any]]:
        try:
            text = self.extract_text_from_xlsx(file_path)
            
            if not text:
                raise ValueError("No data extracted from XLSX")
            
            chunks_with_indices = chunk_text(text, method="smart")
            chunks = [chunk for chunk, _ in chunks_with_indices]
            
            wb = load_workbook(file_path, data_only=True)
            num_sheets = len(wb.sheetnames)
            
            total_rows = 0
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                total_rows += sheet.max_row
            
            wb.close()
            
            metadata = {
                'file_type': 'xlsx',
                'file_path': str(file_path),
                'filename': Path(file_path).name,
                'num_sheets': num_sheets,
                'total_rows': total_rows,
                'text_length': len(text),
                'total_chunks': len(chunks),
                'extension': Path(file_path).suffix
            }
            
            log.info(f"Processed XLSX: {file_path} -> {len(chunks)} chunks")
            return chunks, metadata
            
        except Exception as e:
            log.error(f"Error processing XLSX {file_path}: {e}")
            raise


xlsx_processor = XLSXProcessor()

__all__ = ["XLSXProcessor", "xlsx_processor"]